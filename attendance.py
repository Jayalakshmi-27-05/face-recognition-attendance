from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

file = "attendance.xlsx"

# Create file if not exists
if not os.path.exists(file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Date", "Time"])
    wb.save(file)

def mark_attendance(name):
    wb = load_workbook(file)
    ws = wb.active

    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M:%S")

    # ❗ prevent duplicate entry same day
    for row in ws.iter_rows(values_only=True):
        if row[0] == name and row[1] == date:
            return  # already marked

    ws.append([name, date, time])
    wb.save(file)
